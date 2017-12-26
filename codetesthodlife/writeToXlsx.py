## This file is used for write updated data in excel
import openpyxl
import xlsxwriter
import os
import pandas as pd
from convertBinaryToEmail import binary_to_str
from dataFilesPaths import *

## read excel file
excelFile = openpyxl.load_workbook(input_file_path)
## get workbook to write back
workbook = xlsxwriter.Workbook(updated_input_file_path)

## define sheet name
sheet1 = excelFile.get_sheet_by_name("Sheet1")
## row and col size
row_no = sheet1.max_row
col_no = sheet1.max_column

worksheet = workbook.add_worksheet()
## Sheet cell properties
bold = workbook.add_format({'bold':True, 'font_size':12})
fontsize = workbook.add_format({'font_size':11})
sent_color = workbook.add_format({'font_color': 'green'})
unsent_color = workbook.add_format({'font_color': 'red'})
worksheet.set_column(2, 3, 25)  # set the width of Columns to 25
worksheet.set_column(8, 8, 25)  # set the width of Column 25

def write_to_excel(xl_list_of_dict):
    ## Get All Headers & write in first row
    headers = list(xl_list_of_dict[0].keys())
    for head in range(len(headers)):
        worksheet.write(0, head, headers[head], bold)

    ## Write Remaining values to excel sheet
    for row in range(len(xl_list_of_dict)):
        for col in range(len(xl_list_of_dict[row])):
            values = list(xl_list_of_dict[row].values())
            for val in range(len(values)):
                ## Write status of row
                ## If mail sent --> Green Font
                ## If mail unsent --> Red Font
                if val == 9:
                    if values[val] == 'SENT':
                        worksheet.write(row+1, val, values[val], sent_color)
                    elif values[val] == 'UNSENT':
                        worksheet.write(row+1, val, values[val], unsent_color)
                ## Others Cells in Normal format
                else:
                    worksheet.write(row+1, val, values[val], fontsize)
    workbook.close() # close the file
    return True
